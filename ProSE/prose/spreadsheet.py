"""Read/update the styled Excel workbook.

Design goals (from the spec):
  * 14 columns, fixed left-to-right order.
  * The program only ever fills the 9 DATA columns; the 5 MANUAL columns
    (Status, Pursue, Emailed, Status, Entered SF) are owned by the user and are
    preserved across runs (matched by Solicitation #).
  * De-duplicate by Solicitation #. New solicitations are inserted at the TOP
    (most-recent first); existing ones have their data fields refreshed in place
    without touching the manual columns.
  * Professional styling (borders, alignment, header colour scheme, banding,
    frozen header) is applied automatically and consistently every run, so the
    cosmetics never have to be set up by hand.

Manual-value preservation is done by reading every existing row into memory
keyed by Solicitation #, then rewriting the data area -- this avoids openpyxl's
insert_rows() style-loss pitfalls while keeping the user's typed values.
"""

from __future__ import annotations

import html
import re
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

# Column layout (1-based). Order matches the spec exactly.
HEADERS = [
    "Status",            # 1  manual
    "Solicitation #",    # 2  data
    "Organization",      # 3  data
    "Solicitation Title",  # 4 data
    "Published",         # 5  data
    "Due Date",          # 6  data
    "Pursue",            # 7  manual
    "Emailed",           # 8  manual
    "Status",            # 9  manual (second Status)
    "Entered SF",        # 10 manual
    "Contact Name",      # 11 data
    "Phone",             # 12 data
    "Email",             # 13 data
    "Keyword",           # 14 data (keyword(s) that matched the row)
]

SOLICITATION_COL = 2  # column B holds the dedupe key
DUE_COL = 6           # column F holds the Due Date (used for expiry sorting)

# Due dates arrive as "MM/DD/YYYY hh:mm AM/PM" (HANDS/HiePRO), occasionally as a
# bare "MM/DD/YYYY". A trailing timezone token (HST/HDT) is tolerated.
_DUE_FORMATS = ("%m/%d/%Y %I:%M %p", "%m/%d/%Y %I:%M%p", "%m/%d/%Y")

_TAG_RE = re.compile(r"<[^>]+>")
# Amendment suffix HiePRO appends to a solicitation number ("P27000054
# version: 01"). It changes on every amendment, so it must not be part of the
# dedup key or each amendment appears as a brand-new solicitation.
_VERSION_SUFFIX = re.compile(
    r"\s*[-,]?\s*(?:version|ver|amendment|amend|revision|rev)\s*[:.#]?\s*\d+\s*$",
    re.I,
)


def clean_number(value) -> str:
    """Display form of a Solicitation #: HTML stripped, whitespace collapsed,
    and any trailing amendment suffix removed."""
    text = re.sub(r"\s+", " ", _TAG_RE.sub("", str(value if value is not None else ""))).strip()
    return _VERSION_SUFFIX.sub("", text).strip()


def row_key(value) -> str:
    """Case-insensitive de-duplication key for a Solicitation #."""
    return clean_number(value).upper()


def _parse_due(value) -> datetime | None:
    """Parse a Due Date cell into a datetime, or None if it can't be read.

    A date with no time is treated as the *end* of that day, so a solicitation
    stays active through its whole due date.
    """
    text = re.sub(r"\s+(HST|HDT|HAST|PST|PDT)$", "", str(value or "").strip(), flags=re.I)
    if not text:
        return None
    for fmt in _DUE_FORMATS:
        try:
            due = datetime.strptime(text, fmt)
        except ValueError:
            continue
        if fmt == "%m/%d/%Y":
            due = due.replace(hour=23, minute=59, second=59)
        return due
    return None


def _is_expired(due_value, now: datetime) -> bool:
    """True when ``now`` is past the parsed due date. Unparseable dates are
    treated as NOT expired (never hide a row we can't confidently read)."""
    due = _parse_due(due_value)
    return due is not None and now > due

# Map parser field -> column index.
FIELD_TO_COL = {
    "solicitation_number": 2,
    "organization": 3,
    "title": 4,
    "published": 5,
    "due_date": 6,
    "contact_name": 11,
    "phone": 12,
    "email": 13,
    "keyword": 14,
}

COLUMN_WIDTHS = {
    1: 12, 2: 16, 3: 20, 4: 46, 5: 13, 6: 20,
    7: 10, 8: 10, 9: 12, 10: 12, 11: 20, 12: 16, 13: 28, 14: 22,
}

# --- styling palette --------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")   # deep teal
BAND_FILL = PatternFill("solid", fgColor="EAF1F4")     # light teal banding
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11, color="222222")
# Expired solicitations (past their due date) are struck through and greyed.
EXPIRED_FONT = Font(name="Calibri", size=11, color="9A9A9A", strike=True)
# Cell locking (only enforced by Excel once the sheet is protected). Used to
# make the Solicitation # column read-only while everything else stays editable.
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)
_THIN = Side(style="thin", color="BFC9CE")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
# Same two, but top-aligned. Used ONLY for a cell whose text is taller than its
# row (see _style_data_cell): centring such a cell makes Excel Online/SharePoint
# clip the top AND bottom, showing a band through the middle of the letters,
# while desktop Excel anchors to the top. Top alignment makes both clip only the
# bottom. Cells that fit are centred exactly as before, so nothing else changes.
LEFT_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_WRAP_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
SHEET_TITLE = "Scanned Professional Services"


def _side_sig(side) -> tuple | None:
    """A comparable signature (style, color) for one cell-border side, or None
    when the side has no visible border."""
    if side is None or not side.style:
        return None
    color = getattr(side.color, "rgb", None) if side.color else None
    return (side.style, color)


def _border_sig(border) -> tuple:
    return tuple(_side_sig(s) for s in (border.left, border.right, border.top, border.bottom))


def _detect_border(ws):
    """Return the *dominant* cell border the user has applied to the existing
    grid (header + data), so a rewrite preserves their chosen border weight and
    colour instead of forcing the program's default. ``None`` when the sheet has
    no visibly bordered cells (fresh/blank), in which case callers fall back to
    the default ``BORDER``.

    We reuse the single most common border across all cells rather than a
    per-cell copy, because rows are reordered on every write — a uniform border
    (which is what "a box around every cell" produces) survives that faithfully
    and is applied to newly-added rows too.
    """
    from collections import Counter

    counts: Counter = Counter()
    representative: dict = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            border = ws.cell(row=r, column=c).border
            sig = _border_sig(border)
            if any(sig):  # at least one side is a visible border
                counts[sig] += 1
                # copy() unwraps openpyxl's StyleProxy into a real, assignable
                # Border (a proxy can't be set back onto another cell).
                representative.setdefault(sig, copy(border))
    if not counts:
        return None
    return representative[counts.most_common(1)[0][0]]


def _apply_protection(ws, protect_id: bool) -> None:
    """Turn the Solicitation # column read-only via sheet protection, or clear
    protection entirely. Excel enforces this in its UI; openpyxl (ProSE) writes
    through it regardless, so scans keep refreshing the locked column.

    Structural ops (insert/delete rows & columns) stay blocked, but the things a
    reviewer actually uses — selecting cells, sorting, AutoFilter, formatting —
    are explicitly allowed so the collaborator's workflow isn't hindered. No
    password: this guards against *accidental* edits; Review > Unprotect Sheet
    removes it if ever needed.
    """
    if not protect_id:
        ws.protection = SheetProtection(sheet=False)
        return
    ws.protection = SheetProtection(
        sheet=True,
        selectLockedCells=False, selectUnlockedCells=False,   # allow selecting/copying
        sort=False, autoFilter=False,                         # allow sort + filter
        formatCells=False, formatColumns=False, formatRows=False,  # allow formatting
        insertRows=True, insertColumns=True, insertHyperlinks=True,
        deleteRows=True, deleteColumns=True, pivotTables=True,
        objects=False, scenarios=False,
    )


def _style_header(ws, border=BORDER, protect_id=False) -> None:
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col, value=HEADERS[col - 1])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = border
        if protect_id:
            cell.protection = LOCKED  # headers aren't editable when protected
        # Only seed the default width for columns that have no dimension yet (a
        # fresh sheet, or the newly-added Keyword column). Any column already in
        # column_dimensions keeps its width, preserving the user's adjustments.
        # Note: check membership BEFORE indexing -- ws.column_dimensions[letter]
        # auto-creates an entry, which would defeat the test.
        letter = get_column_letter(col)
        if letter not in ws.column_dimensions:
            ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(col, 16)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _style_data_cell(cell, col: int, banded: bool, expired: bool = False, border=BORDER,
                     protect_id: bool = False, overflows: bool = False) -> None:
    cell.font = EXPIRED_FONT if expired else DATA_FONT
    cell.border = border
    if protect_id:
        # Lock only the Solicitation # column; everything else stays editable.
        cell.protection = LOCKED if col == SOLICITATION_COL else UNLOCKED
    # Title left-wraps; short data/manual columns are centered for a clean look.
    # Phone (12) centers but wraps, so a dual-contact multi-line value shows both.
    # A wrapped cell whose text is taller than the row switches to top alignment
    # so the overflow is clipped at the bottom instead of top-and-bottom.
    if col in (3, 4, 11, 13, 14):
        cell.alignment = LEFT_WRAP_TOP if overflows else LEFT_WRAP
    elif col == 12:
        cell.alignment = CENTER_WRAP_TOP if overflows else CENTER_WRAP
    else:
        cell.alignment = CENTER
    if banded:
        cell.fill = BAND_FILL


# --- row auto-fit -----------------------------------------------------------
# Columns that drive row height: Organization, Contact Name and Email, whichever
# needs the most lines. Deliberately NOT Solicitation Title or Keyword: fitting a
# long title makes every row tall and the sheet reads as empty space. A title
# longer than its row is clipped at the bottom (that cell is top-aligned).
HEIGHT_COLS = (3, 11, 13)
# Every column whose cells wrap — the candidates for an overflow check.
WRAP_COLS = (3, 4, 11, 12, 13, 14)
LINE_POINTS = 15.0      # height of one line of 11pt Calibri
MAX_ROW_HEIGHT = 120.0  # ~8 lines, so one pathological cell can't eat the screen


def _column_chars(ws, col: int) -> int:
    """Roughly how many characters fit on one line of a column, from its *live*
    width (the user may have widened it, and that is preserved across scans)."""
    letter = get_column_letter(col)
    width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
    if not width:
        width = COLUMN_WIDTHS.get(col, 16)
    return max(4, int(float(width)) - 1)  # one char per width unit, less padding


def _wrapped_lines(value, chars: int) -> int:
    """Lines a wrapped cell needs: explicit newlines plus greedy word wrapping."""
    total = 0
    for segment in str(value if value is not None else "").split("\n"):
        words = segment.split()
        if not words:
            total += 1
            continue
        lines, used = 1, 0
        for word in words:
            needed = len(word) if used == 0 else len(word) + 1
            if used + needed <= chars:
                used += needed
                continue
            if used:      # start a new line; a word at column 0 already has one
                lines += 1
            used = len(word)
            while used > chars:  # a single token wider than the column
                lines += 1
                used -= chars
        total += lines
    return total


def _autofit_row(ws, row: int) -> float:
    """Set (and return) a row's height to fit its Organization / Contact Name /
    Email cells — the equivalent of double-clicking the row's bottom border in
    Excel, but measured only against the columns worth growing for.

    Excel only auto-fits wrapped text when it recalculates a row, which it does
    not do for rows written by a library, so ProSE computes the height itself.
    Needed now that Contact Name / Phone / Email can hold two contacts.
    """
    lines = 1
    for col in HEIGHT_COLS:
        lines = max(lines, _wrapped_lines(ws.cell(row=row, column=col).value,
                                          _column_chars(ws, col)))
    height = min(MAX_ROW_HEIGHT, lines * LINE_POINTS)
    ws.row_dimensions[row].height = height
    return height


def _overflowing_cols(ws, row: int, height: float) -> set[int]:
    """Wrapped columns in this row whose text is taller than the row.

    Normally only Solicitation Title (never measured for height) and, on a capped
    row, whichever column hit the cap. These are the cells that get top-aligned.
    """
    return {
        col for col in WRAP_COLS
        if _wrapped_lines(ws.cell(row=row, column=col).value,
                          _column_chars(ws, col)) * LINE_POINTS > height + 0.01
    }


# Role tag appended to a contact name when a notice lists two distinct people.
_ROLE_TAG = re.compile(r"\s*\((?:Specifications|Buyer)\)\s*$", re.I)

# An HTML entity left in a cell by a scan that predates the unescape fix.
_ENTITY_RE = re.compile(r"&(?:#\d+|#x[0-9A-Fa-f]+|[A-Za-z][A-Za-z0-9]{1,31});")


def _decode_entities(rec: dict) -> bool:
    """Decode HTML entities (``&#x27;`` → ``'``) left in ProSE-owned cells.

    Only the data columns are touched — never the user's manual columns, where a
    literal ``&amp;`` might be intentional. Rows for closed solicitations are
    never refreshed by a scan, so without this they keep the raw entities.
    """
    changed = False
    for col in FIELD_TO_COL.values():
        value = rec.get(col)
        if isinstance(value, str) and _ENTITY_RE.search(value):
            rec[col] = html.unescape(value)
            changed = True
    return changed


def _collapse_repeated_contact(rec: dict) -> bool:
    """Rewrite a stored two-line contact that is the same person in both roles.

    Rows for solicitations that have since closed are never refreshed by a scan,
    so without this they would keep the doubled contact (and the tall row) they
    were written with. Only collapses when name, phone and email all agree on
    both lines — anything genuinely different is left alone.
    """
    name = str(rec.get(11) or "")
    if "\n" not in name:
        return False
    names = [_ROLE_TAG.sub("", part).strip() for part in name.split("\n")]
    if len(names) != 2 or names[0].lower() != names[1].lower():
        return False

    def parts(col: int) -> list[str]:
        return [p.strip() for p in str(rec.get(col) or "").split("\n")]

    phones, emails = parts(12), parts(13)
    for values in (phones, emails):
        if len(values) > 1 and len({v.lower() for v in values}) > 1:
            return False  # same name, different phone/email — keep both lines
    rec[11] = names[0]
    rec[12] = phones[0]
    rec[13] = emails[0]
    return True


def _office_owner_file(path: Path) -> Path:
    """Excel's owner/lock file for an open workbook: ``~$<name>.xlsx`` beside it."""
    return path.with_name("~$" + path.name)


def _held_by_process(path: Path) -> bool:
    """True when some process still holds ``path`` open.

    Windows denies sharing on both the workbook and Excel's owner file while a
    workbook is open, so a failed open is a reliable liveness signal — and a
    *successful* one proves nothing is holding the file, whatever its age.
    """
    try:
        with open(path, "rb+"):
            return False
    except PermissionError:
        return True
    except OSError:
        return False  # missing, or unreadable for an unrelated reason


# Lock states returned by excel_lock_state().
LOCK_FREE = "free"    # nothing holds the workbook
LOCK_OPEN = "open"    # Excel (or another process) has it open right now
LOCK_STALE = "stale"  # an owner file no process holds — an Excel crash leftover


def excel_lock_state(path) -> str:
    """Classify the workbook's lock as free / open / stale.

    A write-permission probe alone is not enough for a OneDrive/SharePoint file:
    with AutoSave on, Excel co-authors through the sync client and does **not**
    take an exclusive OS lock, so a save would quietly succeed and then race the
    open copy. Excel does still drop its owner file next to the workbook, which
    is the dependable signal in that setup.

    Both signals are checked, so a live session is caught even if a future Excel
    build shares its owner file more permissively. ``stale`` is only returned
    when *neither* file is held by anything, which is exactly the state an Excel
    crash leaves behind.
    """
    path = Path(path)
    owner = _office_owner_file(path)
    try:
        has_owner = owner.exists()
    except OSError:
        return LOCK_FREE
    if _held_by_process(path) or (has_owner and _held_by_process(owner)):
        return LOCK_OPEN
    return LOCK_STALE if has_owner else LOCK_FREE


def clear_stale_lock(path) -> bool:
    """Delete an orphaned ``~$`` owner file, returning True if one was removed.

    Failsafe for the one way the lock guard could wedge permanently: if Excel
    crashes (or is killed) it never cleans up its owner file, and every later
    scan would skip forever. Only ever deletes a file that no process holds, so
    it cannot pull the guard out from under a live Excel session.
    """
    if excel_lock_state(path) != LOCK_STALE:
        return False
    try:
        _office_owner_file(Path(path)).unlink()
    except OSError:
        return False
    return True


def is_open_in_excel(path) -> bool:
    """True when the workbook is open right now (a stale owner file is not)."""
    return excel_lock_state(path) == LOCK_OPEN


def _new_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE
    _style_header(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return wb


def _read_rows(ws) -> list[dict]:
    """Read existing data rows into a list of {col_index: value} dicts."""
    rows = []
    for r in range(2, ws.max_row + 1):
        record = {c: ws.cell(row=r, column=c).value for c in range(1, len(HEADERS) + 1)}
        # Skip completely empty rows.
        if all(v in (None, "") for v in record.values()):
            continue
        rows.append(record)
    return rows


def update_spreadsheet(path, details: list[dict], skip_on_lock: bool = False,
                       protect_id: bool = False) -> dict:
    """Merge scanned ``details`` into the workbook at ``path``.

    Returns {"new": int, "updated": int, "total_rows": int, ...}.

    ``skip_on_lock`` controls what happens when the file is locked (open in
    Excel / held by a co-author): when False (local use) the merge is saved to a
    timestamped sibling so results aren't lost; when True (a shared
    SharePoint/OneDrive workbook) the write is SKIPPED instead — no sibling is
    dropped into the shared library, and the next scheduled scan merges once the
    file is closed. Scans re-query the source every run, so nothing is lost
    permanently either way.

    ``protect_id`` locks the Solicitation # column (via Excel sheet protection)
    so a collaborator can't accidentally edit the key ProSE matches rows on;
    every other column stays editable. Re-applied every run.
    """
    path = Path(path)
    # Always clear a crash-orphaned owner file first, in both local and shared
    # mode: it is never valid, and left alone it would wedge every future scan.
    stale_lock_cleared = clear_stale_lock(path)
    if skip_on_lock and is_open_in_excel(path):
        # It really is open (possibly co-authoring, which takes no OS lock).
        # Skip rather than fight that copy; the next scan merges.
        return {
            "new": 0, "updated": 0, "total_rows": 0, "saved_to": str(path),
            "diverted": False, "skipped_locked": True, "duplicates_removed": 0,
            "stale_lock_cleared": stale_lock_cleared,
        }
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = _new_workbook(path)
        ws = wb.active

    existing = _read_rows(ws)
    # Capture the border the user has applied before we wipe the data area, so
    # the rewrite preserves it (falls back to the default on a fresh sheet).
    data_border = _detect_border(ws) or BORDER

    # Index existing rows by solicitation # (normalised: HTML stripped, any
    # "version: NN" amendment suffix removed, case-folded). Rows without a
    # solicitation # are carried through untouched at the end.
    by_key: dict[str, dict] = {}
    carry: list[dict] = []
    order: list[str] = []
    duplicates_removed = 0
    contacts_collapsed = 0
    entities_decoded = 0
    for rec in existing:
        if _decode_entities(rec):
            entities_decoded += 1
        if _collapse_repeated_contact(rec):
            contacts_collapsed += 1
        number = clean_number(rec.get(SOLICITATION_COL))
        key = row_key(number)
        if not key:
            carry.append(rec)
            continue
        rec[SOLICITATION_COL] = number  # self-heal a stored "… version: 01"
        kept = by_key.get(key)
        if kept is None:
            by_key[key] = rec
            order.append(key)
            continue
        # The same solicitation is stored twice: an amendment row left by an
        # older ProSE build, or a hand-pasted copy. Keep the first occurrence
        # (its position) and fold in any value only the duplicate carries, so
        # manual columns typed on either copy survive the collapse.
        for col, value in rec.items():
            if value not in (None, "") and kept.get(col) in (None, ""):
                kept[col] = value
        duplicates_removed += 1

    new_count = 0
    updated_count = 0
    new_keys: list[str] = []

    def published_sort_key(detail: dict):
        # newest first; expects MM/DD/YYYY, falls back to string
        val = detail.get("published") or ""
        parts = val.split("/")
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            mm, dd, yy = parts
            return (int(yy), int(mm), int(dd))
        return (0, 0, 0)

    for detail in sorted(details, key=published_sort_key):  # ascending; prepended -> newest ends on top
        number = clean_number(detail.get("solicitation_number"))
        key = row_key(number)
        if not key:
            continue
        if key in by_key:
            rec = by_key[key]
            for field, col in FIELD_TO_COL.items():
                rec[col] = detail.get(field, "")
            updated_count += 1
        else:
            rec = {c: "" for c in range(1, len(HEADERS) + 1)}
            for field, col in FIELD_TO_COL.items():
                rec[col] = detail.get(field, "")
            by_key[key] = rec
            new_keys.append(key)
            new_count += 1
        rec[SOLICITATION_COL] = number  # store the clean number, never a variant

    # Final order: newest new-keys on top (new_keys is oldest->newest, so reverse),
    # then previously-existing rows in their prior order, then carry rows.
    final_order = list(reversed(new_keys)) + [k for k in order if k not in new_keys]
    ordered_records = [by_key[k] for k in final_order] + carry

    # Sink expired solicitations (past their due date) to the bottom, preserving
    # the relative order within each group. Recomputed every run against "now".
    now = datetime.now()
    active = [rec for rec in ordered_records if not _is_expired(rec.get(DUE_COL), now)]
    expired = [rec for rec in ordered_records if _is_expired(rec.get(DUE_COL), now)]
    ordered_records = active + expired
    first_expired = len(active)  # rows at this index and below are struck through

    # Wipe existing data area (values + styles) then rewrite.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for i, rec in enumerate(ordered_records):
        row = i + 2
        banded = (i % 2 == 1)
        is_expired = i >= first_expired
        # Values first: the row height (and so which cells overflow it) can only
        # be measured once the text is in place.
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col, value=rec.get(col))
        height = _autofit_row(ws, row)
        overflowing = _overflowing_cols(ws, row, height)
        for col in range(1, len(HEADERS) + 1):
            _style_data_cell(ws.cell(row=row, column=col), col, banded, is_expired,
                             data_border, protect_id, col in overflowing)

    # Drop heights left behind by a previously longer sheet.
    last_row = len(ordered_records) + 1
    for stale in [r for r in ws.row_dimensions if r > last_row]:
        del ws.row_dimensions[stale]

    # Re-assert header styling (cheap, keeps look consistent).
    _style_header(ws, data_border, protect_id)
    # Lock the Solicitation # column (or clear protection) every run.
    _apply_protection(ws, protect_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    saved_to = path
    try:
        wb.save(path)
    except PermissionError:
        # The workbook is open in Excel / held by a co-author (Windows locks it).
        if skip_on_lock:
            # Shared library: skip this run rather than littering a conflicting
            # sibling into the synced folder. The next scan will merge.
            return {
                "new": 0,
                "updated": 0,
                "total_rows": len(ordered_records),
                "saved_to": str(path),
                "diverted": False,
                "skipped_locked": True,
                "duplicates_removed": 0,  # nothing was written
                "stale_lock_cleared": stale_lock_cleared,
            }
        # Local use: don't lose the scan — save to a timestamped sibling.
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        saved_to = path.with_name(f"{path.stem} (scan {stamp}){path.suffix}")
        wb.save(saved_to)

    return {
        "new": new_count,
        "updated": updated_count,
        "total_rows": len(ordered_records),
        "saved_to": str(saved_to),
        "diverted": saved_to != path,
        "skipped_locked": False,
        "duplicates_removed": duplicates_removed,
        "contacts_collapsed": contacts_collapsed,
        "entities_decoded": entities_decoded,
        "stale_lock_cleared": stale_lock_cleared,
    }
